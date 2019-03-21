import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook

def main():
    wb = Workbook()
    sheet1 = wb.active
    file_name = "NaverDaumRank.xlsx"
    naverRankedList = []
    daumRankedList = []
    url = "http://www.naver.com"
    req = urllib.request.Request(url)
    sc  = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sc, "html.parser")
    for data in soup.find_all("span", class_ = "ah_k"):
        naverRankedList.append(data.get_text())
    print(naverRankedList)
    url = "http://www.daum.net"
    req = urllib.request.Request(url)
    sc  = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sc, "html.parser")
    for data in soup.select('a.link_issue'):
        daumRankedList.append(data.text)
    print(daumRankedList)
    real_daum_list = daumRankedList[: :2]
    print(real_daum_list)
    for index in range(1,  21):
        sheet1.cell(row = index, column = 1).value = index
        sheet1.cell(row = index, column = 2).value = naverRankedList[index - 1]
    for index in range(1,  len(real_daum_list) + 1):
        sheet1.cell(row = index, column = 3).value = index
        sheet1.cell(row = index, column = 4).value = real_daum_list[index - 1]
    wb.save(filename = "190321_12_51.xlsx")
if __name__ == "__main__":
    main()